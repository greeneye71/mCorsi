from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from ..extensions import db
from ..models import (
    Enrollment,
    ImportBatch,
    ImportRow,
    ParticipantProfile,
    Role,
    User,
    normalize_email,
)
from .courses import create_course
from .storage import path_for


class HistoricalImportError(ValueError):
    pass


MAX_IMPORT_ROWS = 10_000
MAX_IMPORT_COLUMNS = 100
ATTENDANCE_STATUSES = {"pending", "attended", "absent"}


def _normal_header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"\s+", " ", text)
    return text


def _valid_email(value: Any) -> str:
    text = str(value or "").strip().casefold()
    if not text or text == "anonymous":
        return ""
    try:
        return validate_email(text, check_deliverability=False).normalized.casefold()
    except EmailNotValidError:
        return ""


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _worksheet_matrix(path: Path) -> tuple[list[Any], list[tuple[Any, ...]]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise HistoricalImportError("Il file Excel non è leggibile o è danneggiato.") from exc
    try:
        for worksheet in workbook.worksheets:
            min_col, min_row, max_col, max_row = range_boundaries(
                worksheet.calculate_dimension()
            )
            row_count = max_row - min_row + 1
            column_count = max_col - min_col + 1
            if row_count > MAX_IMPORT_ROWS or column_count > MAX_IMPORT_COLUMNS:
                raise HistoricalImportError(
                    f"Il file supera il limite di {MAX_IMPORT_ROWS} righe o "
                    f"{MAX_IMPORT_COLUMNS} colonne."
                )
            values = list(
                worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )
            if len(values) >= 2 and any(value is not None for value in values[0]):
                return list(values[0]), values[1:]
    finally:
        workbook.close()
    raise HistoricalImportError("Il file non contiene una tabella con risposte.")


def _best_column(headers: list[Any], rows: list[tuple[Any, ...]], kind: str) -> int | None:
    aliases = {
        "email": {"email", "e-mail", "posta elettronica", "indirizzo email"},
        "first_name": {"nome", "nome2", "nome partecipante"},
        "last_name": {"cognome"},
        "birth_place": {"luogo di nascita"},
        "birth_date": {"data di nascita"},
        "certificate_title": {"titolo da riportare sull'attestato (sig. / dott. ecc....)", "titolo"},
    }
    candidates = []
    for index, header in enumerate(headers):
        normalized = _normal_header(header)
        if normalized.startswith("points -") or normalized.startswith("feedback -"):
            continue
        if normalized in aliases[kind]:
            if kind == "email":
                score = sum(bool(_valid_email(row[index] if index < len(row) else None)) for row in rows)
            else:
                score = sum(bool(_cell_text(row[index] if index < len(row) else None)) for row in rows)
            candidates.append((score, index))
    return max(candidates, default=(0, None))[1]


def parse_historical_workbook(path: Path) -> tuple[list[dict], dict]:
    headers, rows = _worksheet_matrix(path)
    mapping = {
        key: _best_column(headers, rows, key)
        for key in ("email", "first_name", "last_name", "birth_place", "birth_date", "certificate_title")
    }
    required = [key for key in ("email", "first_name", "last_name") if mapping[key] is None]
    if required:
        raise HistoricalImportError(
            "Non riconosco le colonne obbligatorie: " + ", ".join(required) + "."
        )

    candidates = []
    for row_number, row in enumerate(rows, start=2):
        def value(key):
            index = mapping[key]
            return row[index] if index is not None and index < len(row) else None

        first_name = _cell_text(value("first_name"))
        last_name = _cell_text(value("last_name"))
        birth_date = _date_value(value("birth_date"))
        email_raw = _cell_text(value("email")).casefold()
        email = _valid_email(email_raw)
        if not any((first_name, last_name, email)):
            continue
        identity = (
            f"person:{first_name.casefold()}|{last_name.casefold()}|{birth_date.isoformat() if birth_date else ''}"
            if first_name and last_name
            else f"email:{email or email_raw}"
        )
        candidates.append(
            {
                "identity": identity,
                "row_number": row_number,
                "email": email,
                "email_raw": email_raw,
                "first_name": first_name,
                "last_name": last_name,
                "birth_place": _cell_text(value("birth_place")),
                "birth_date": birth_date,
                "certificate_title": _cell_text(value("certificate_title")),
            }
        )

    grouped = defaultdict(list)
    for candidate in candidates:
        grouped[candidate["identity"]].append(candidate)
    parsed = []
    for group in grouped.values():
        representative = group[-1].copy()
        valid_emails = [item["email"] for item in group if item["email"]]
        representative["email"] = Counter(valid_emails).most_common(1)[0][0] if valid_emails else ""
        representative["source_rows"] = [item["row_number"] for item in group]
        warnings = []
        if len(group) > 1:
            warnings.append(f"{len(group)} risposte consolidate")
        raw_emails = {item["email_raw"] for item in group if item["email_raw"]}
        if len(raw_emails) > 1:
            warnings.append("email discordanti; scelta quella più frequente")
        if not representative["email"]:
            warnings.append("email mancante o non valida")
        representative["warning"] = "; ".join(warnings)
        parsed.append(representative)
    parsed.sort(key=lambda item: (item["last_name"].casefold(), item["first_name"].casefold()))
    mapping_labels = {
        key: str(headers[index]) if index is not None else ""
        for key, index in mapping.items()
    }
    return parsed, mapping_labels


def prepare_batch(batch: ImportBatch) -> ImportBatch:
    parsed, mapping = parse_historical_workbook(path_for(batch.stored_file))
    if not parsed:
        raise HistoricalImportError("Non è stato trovato alcun partecipante.")
    batch.detected_mapping = mapping
    for index, item in enumerate(parsed, start=1):
        batch.rows.append(
            ImportRow(
                row_number=index,
                source_rows=item["source_rows"],
                email=item["email"],
                first_name=item["first_name"],
                last_name=item["last_name"],
                birth_place=item["birth_place"],
                birth_date=item["birth_date"],
                certificate_title=item["certificate_title"],
                status="ready" if item["email"] else "error",
                warning=item["warning"],
            )
        )
    batch.summary = {
        "source_responses": sum(len(item["source_rows"]) for item in parsed),
        "people": len(parsed),
        "ready": sum(bool(item["email"]) for item in parsed),
    }
    db.session.flush()
    return batch


def _find_participant(row: ImportRow) -> User | None:
    existing = User.query.filter_by(email=normalize_email(row.email)).first()
    if existing:
        if existing.has_role("participant"):
            return existing
        raise HistoricalImportError("L'email appartiene a un operatore e non può essere importata.")
    candidates = (
        User.query.options(joinedload(User.participant_profile))
        .filter(
            User.roles.any(name="participant"),
            func.lower(User.first_name) == row.first_name.casefold(),
            func.lower(User.last_name) == row.last_name.casefold(),
        )
        .all()
    )
    for user in candidates:
        profile = user.participant_profile
        if (
            user.first_name.casefold() == row.first_name.casefold()
            and user.last_name.casefold() == row.last_name.casefold()
            and profile
            and profile.birth_date == row.birth_date
        ):
            return user
    return None


def confirm_batch(
    batch: ImportBatch, *, actor: User, attendance_status: str
) -> ImportBatch:
    if batch.status != "preview":
        raise HistoricalImportError("Questo import è già stato elaborato.")
    if attendance_status not in ATTENDANCE_STATUSES:
        raise HistoricalImportError("Lo stato delle presenze importate non è valido.")
    course = create_course(
        actor=actor,
        data={
            "title": batch.course_title,
            "description": f"Corso storico importato da {batch.stored_file.original_name}.",
            "status": "completed",
            "is_historical": True,
            "referent_user_id": actor.id,
            "session_date": batch.course_date,
            "start_time": time(9, 0),
            "end_time": time(13, 0),
            "delivery_mode": "online",
            "meeting_url": "",
            "certificate_validity_months": None,
        },
    )
    participant_role = Role.query.filter_by(name="participant").one()
    imported = skipped = errors = 0
    for row in batch.rows:
        if row.status != "ready":
            errors += 1
            continue
        try:
            participant = _find_participant(row)
            if participant is None:
                participant = User(
                    email=normalize_email(row.email),
                    first_name=row.first_name,
                    last_name=row.last_name,
                    profile_completed=bool(row.first_name and row.last_name and row.birth_place and row.birth_date),
                )
                participant.roles.append(participant_role)
                participant.participant_profile = ParticipantProfile(
                    birth_place=row.birth_place,
                    birth_date=row.birth_date,
                    certificate_title=row.certificate_title,
                )
                db.session.add(participant)
                db.session.flush()
            else:
                participant.first_name = participant.first_name or row.first_name
                participant.last_name = participant.last_name or row.last_name
                if participant.participant_profile is None:
                    participant.participant_profile = ParticipantProfile()
                profile = participant.participant_profile
                profile.birth_place = profile.birth_place or row.birth_place
                profile.birth_date = profile.birth_date or row.birth_date
                profile.certificate_title = profile.certificate_title or row.certificate_title
                participant.profile_completed = bool(
                    participant.first_name and participant.last_name and profile.birth_place and profile.birth_date
                )
            if Enrollment.query.filter_by(course_id=course.id, participant_user_id=participant.id).first():
                row.status = "skipped"
                row.warning = (row.warning + "; " if row.warning else "") + "già iscritto"
                skipped += 1
                continue
            enrollment = Enrollment(
                course=course,
                participant=participant,
                attendance_status=attendance_status,
            )
            db.session.add(enrollment)
            db.session.flush()
            row.participant_user_id = participant.id
            row.enrollment_id = enrollment.id
            row.status = "imported"
            imported += 1
        except HistoricalImportError as exc:
            row.status = "error"
            row.warning = str(exc)
            errors += 1
        except Exception as exc:
            row.status = "error"
            row.warning = f"errore di importazione: {type(exc).__name__}"
            errors += 1
    batch.course = course
    batch.status = "completed" if not errors else "completed_with_errors"
    batch.completed_at = datetime.now(timezone.utc)
    batch.summary = {
        **batch.summary,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "attendance_status": attendance_status,
    }
    db.session.flush()
    return batch
